# Copyright 2026 Ashita Aggarwal and Suraj Commuri
# SPDX-License-Identifier: Apache-2.0

"""A checkpoint as a spreadsheet, and a spreadsheet back into decisions.

A checkpoint with a hundred codes is a codebook, and researchers keep
codebooks in spreadsheets — with co-coders, with track changes, in the tool
they trust. So a code-review checkpoint (and the literature-synthesis
extraction table) can be exported as an .xlsx workbook, edited anywhere that
opens one, and uploaded again. The upload is parsed into exactly the
decisions the review screen's buttons produce; nothing is applied here. The
screen loads them as staged edits, the researcher reads them, and presses
Approve & continue — one apply path, one audit trail.

Rules (every one refuses or reports rather than guesses):
  * the workbook must belong to this checkpoint (a workbook-defined name and
    the About sheet both carry the checkpoint id);
  * rows are matched by id; an unknown id is ignored and listed; a code with
    no row is unchanged — only the *delete* action deletes;
  * a changed name or definition is an edit, no action word needed; a blank
    name keeps the old one; a blank definition is a deliberate clearing;
  * *merge into* takes an id or a name that matches exactly one code;
  * a row with no id and a name is a new code;
  * notes travel with the decision into the audit trail.
"""

from __future__ import annotations

import hashlib
import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

MARK = "qualilens_checkpoint"          # workbook-defined name holding the checkpoint id
MAX_BYTES = 10 * 1024 * 1024
MAX_ROWS = 50_000
SUPPORTED = ("code_review", "extraction_review")

CODE_COLUMNS = [                       # (header, width, editable-by-researcher)
    ("id", 16, False), ("name", 34, True), ("definition", 60, True),
    ("excerpts", 10, False), ("sample excerpt", 60, False),
    ("action", 12, True), ("merge into", 30, True), ("notes", 40, True),
]
PAPER_FIXED = [("source_id", 16, False), ("filename", 28, False),
               ("label", 22, True), ("citation", 48, True)]
PAPER_TAIL = [("cited work", 40, True), ("exclude", 10, True), ("notes", 40, True)]

_HEAD_FONT = Font(bold=True)
_HEAD_FILL = PatternFill("solid", fgColor="ECECEF")
_WRAP = Alignment(wrap_text=True, vertical="top")
_TEXT = "@"                             # Excel's text number format: typed input stays text


class SheetError(ValueError):
    pass


# ---------------------------------------------------------------- export

def _sheet(wb: Workbook, title: str, columns: list, first: bool = False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    for i, (head, width, editable) in enumerate(columns, 1):
        c = ws.cell(row=1, column=i, value=head)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    return ws


def _put(ws, row: int, col: int, value: Any, editable: bool = False):
    c = ws.cell(row=row, column=col, value="" if value is None else str(value))
    c.alignment = _WRAP
    c.number_format = _TEXT
    return c


def _about(wb: Workbook, kind: str, meta: dict, rules: list[str]):
    ws = wb.create_sheet("About")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    rows = [("project", meta.get("project_name", "")), ("run id", meta.get("run_id", "")),
            ("checkpoint id", meta.get("checkpoint_id", "")), ("checkpoint", meta.get("title", "")),
            ("stage", meta.get("stage", "")), ("kind", kind), ("exported at", meta.get("exported_at", "")),
            ("", ""), ("how to use this file", "")]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = _HEAD_FONT
        _put(ws, i, 2, v)
    for j, line in enumerate(rules, len(rows) + 1):
        _put(ws, j, 2, line)
    wb.defined_names[MARK] = DefinedName(MARK, attr_text=f'"{meta.get("checkpoint_id", "")}"')


CODE_RULES = [
    "Edit a name or a definition in place; that is a rename or a redefinition. No action word is needed.",
    "Leave 'action' blank to keep a code. Choose 'merge' and put the target's id or exact name in 'merge into'; choose 'delete' to delete.",
    "A row you remove from this sheet leaves that code unchanged. Only the 'delete' action deletes.",
    "Add a code by adding a row with an empty id and a name.",
    "Do not edit the id column. 'excerpts' and 'sample excerpt' are for orientation and are not read back.",
    "'notes' travel with your decision into the run's audit trail.",
    "Upload the file on the same checkpoint. Its decisions load into the screen for you to check before you approve.",
]
PAPER_RULES = [
    "Edit the label, the citation, or any of the five fields in place. A blank label keeps the old one.",
    "Put 'yes' in 'exclude' to leave a paper out of the synthesis; blank or 'no' keeps it in. Its extraction stays in the record.",
    "A row you remove from this sheet leaves that paper unchanged.",
    "Do not edit source_id or filename. The Quotes sheet is for orientation and is not read back.",
    "'notes' travel with your edits into the run's audit trail.",
    "Upload the file on the same checkpoint. Its edits load into the screen for you to check before you approve.",
]


def export_workbook(kind: str, payload: dict, meta: dict, excerpts: list[dict] | None = None) -> bytes:
    """meta: project_name, run_id, checkpoint_id, title, stage, exported_at.
    excerpts (code_review): [{code_id, code, via, source, quote, memo}]."""
    if kind not in SUPPORTED:
        raise SheetError("This checkpoint has no spreadsheet form.")
    wb = Workbook()
    if kind == "code_review":
        ws = _sheet(wb, "Codes", CODE_COLUMNS, first=True)
        items = payload.get("items") or []
        for r, it in enumerate(items, 2):
            sample = (it.get("sample_excerpts") or [{}])[0].get("quote", "") if it.get("sample_excerpts") else ""
            vals = [it.get("id", ""), it.get("name", ""), it.get("definition", ""),
                    it.get("excerpt_count", 0), sample, "", "", ""]
            for c, v in enumerate(vals, 1):
                _put(ws, r, c, v)
        dv = DataValidation(type="list", formula1='"keep,merge,delete"', allow_blank=True,
                            showErrorMessage=True, errorTitle="Action",
                            error="Use keep, merge, or delete (or leave blank to keep).")
        ws.add_data_validation(dv)
        dv.add(f"F2:F{max(2, len(items) + 1) + 200}")   # room for added rows too
        ex = _sheet(wb, "Excerpts", [("code id", 16, False), ("code", 30, False), ("via", 26, False),
                                     ("source", 28, False), ("quote", 80, False), ("memo", 40, False)])
        for r, e in enumerate(excerpts or [], 2):
            for c, k in enumerate(("code_id", "code", "via", "source", "quote", "memo"), 1):
                _put(ex, r, c, e.get(k, ""))
        _about(wb, kind, meta, CODE_RULES)
    else:
        fields = payload.get("fields") or []
        labels = payload.get("field_labels") or {}
        cols = PAPER_FIXED + [(labels.get(k, k), 48, True) for k in fields] + PAPER_TAIL
        ws = _sheet(wb, "Papers", cols, first=True)
        rows = payload.get("rows") or []
        for r, row in enumerate(rows, 2):
            vals = [row.get("source_id", ""), row.get("filename", ""), row.get("label", ""),
                    row.get("citation", "")]
            vals += [(row.get("fields") or {}).get(k, "") for k in fields]
            vals += [row.get("cited_work", ""), "yes" if row.get("excluded") else "", ""]
            for c, v in enumerate(vals, 1):
                _put(ws, r, c, v)
        dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
        ws.add_data_validation(dv)
        excl_col = get_column_letter(len(cols) - 1)
        dv.add(f"{excl_col}2:{excl_col}{max(2, len(rows) + 1)}")
        q = _sheet(wb, "Quotes", [("source_id", 16, False), ("paper", 24, False), ("field", 14, False),
                                  ("quote", 90, False), ("memo", 40, False)])
        for r, e in enumerate(excerpts or [], 2):
            for c, k in enumerate(("source_id", "paper", "field", "quote", "memo"), 1):
                _put(q, r, c, e.get(k, ""))
        _about(wb, kind, meta, PAPER_RULES)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- import

def _norm(v: Any) -> str:
    if v is None:
        return ""
    s = v if isinstance(v, str) else str(v)
    return s.replace("\r\n", "\n").strip()


def _open(data: bytes):
    if len(data) > MAX_BYTES:
        raise SheetError(f"That file is larger than {MAX_BYTES // (1024 * 1024)} MB; a checkpoint "
                         "workbook is far smaller. Refusing.")
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — anything but a workbook is refused
        raise SheetError("That file is not an Excel workbook (.xlsx). Numbers and Google Sheets "
                         "can export one.") from e


def _marked_checkpoint(wb) -> tuple[str, dict]:
    """The checkpoint id the workbook claims, from the defined name or the
    About sheet, plus what About says (for a useful refusal)."""
    about: dict = {}
    if "About" in wb.sheetnames:
        for row in wb["About"].iter_rows(min_row=1, max_row=12, max_col=2, values_only=True):
            if row and row[0]:
                about[_norm(row[0]).casefold()] = _norm(row[1] if len(row) > 1 else "")
    marked = ""
    try:
        dn = wb.defined_names.get(MARK) if hasattr(wb.defined_names, "get") else None
        if dn is not None and dn.attr_text:
            marked = dn.attr_text.strip().strip('"')
    except Exception:  # noqa: BLE001
        marked = ""
    return marked or about.get("checkpoint id", ""), about


def _headers(ws) -> dict[str, int]:
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    return {_norm(h).casefold(): i for i, h in enumerate(first) if _norm(h)}


def _rows(ws):
    n = 0
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        n += 1
        if n > MAX_ROWS:
            raise SheetError(f"More than {MAX_ROWS} rows; refusing.")
        if row and any(_norm(v) for v in row):
            yield r, row


def _cell(row, headers: dict, name: str) -> str:
    i = headers.get(name)
    return _norm(row[i]) if i is not None and i < len(row) else ""


def parse_workbook(kind: str, payload: dict, data: bytes, checkpoint_id: str) -> dict:
    if kind not in SUPPORTED:
        raise SheetError("This checkpoint has no spreadsheet form.")
    wb = _open(data)
    marked, about = _marked_checkpoint(wb)
    if not marked:
        raise SheetError("That workbook was not exported from a QualiLens checkpoint "
                         "(no checkpoint mark). Download the sheet from this checkpoint and edit that.")
    if marked != checkpoint_id:
        where = about.get("checkpoint", "another checkpoint")
        proj = about.get("project", "")
        raise SheetError(f"That workbook belongs to a different checkpoint — “{where}”"
                         f"{' of project ' + repr(proj) if proj else ''}. Download this "
                         "checkpoint's own sheet.")
    out = _parse_codes(wb, payload) if kind == "code_review" else _parse_papers(wb, payload)
    out["sha256"] = hashlib.sha256(data).hexdigest()
    return out


def _parse_codes(wb, payload: dict) -> dict:
    if "Codes" not in wb.sheetnames:
        raise SheetError("The workbook has no 'Codes' sheet.")
    ws = wb["Codes"]
    h = _headers(ws)
    for need in ("id", "name"):
        if need not in h:
            raise SheetError(f"The Codes sheet has no '{need}' column. Keep the exported columns.")
    items = {it["id"]: it for it in (payload.get("items") or []) if it.get("id")}
    by_name: dict[str, list[str]] = {}
    for it in items.values():
        by_name.setdefault(_norm(it.get("name")).casefold(), []).append(it["id"])
    decisions: list[dict] = []
    additions: list[dict] = []
    ignored: list[dict] = []
    summary = {"renamed": 0, "redefined": 0, "merged": 0, "deleted": 0, "added": 0,
               "unchanged": 0, "ignored": 0, "with_notes": 0}
    seen: set[str] = set()
    for rownum, row in _rows(ws):
        cid = _cell(row, h, "id")
        name = _cell(row, h, "name")
        definition_i = h.get("definition")
        definition = _norm(row[definition_i]) if definition_i is not None and definition_i < len(row) else None
        action = _cell(row, h, "action").casefold()
        target = _cell(row, h, "merge into")
        notes = _cell(row, h, "notes")
        if not cid:
            if name:
                add = {"name": name, "definition": definition or ""}
                if notes:
                    add["notes"] = notes
                    summary["with_notes"] += 1
                additions.append(add)
                summary["added"] += 1
            continue
        if cid not in items:
            ignored.append({"row": rownum, "id": cid, "reason": "not a code of this checkpoint"})
            continue
        if cid in seen:
            ignored.append({"row": rownum, "id": cid, "reason": "duplicate row for this id; first row used"})
            continue
        seen.add(cid)
        it = items[cid]
        if action not in ("", "keep", "rename", "merge", "delete"):
            ignored.append({"row": rownum, "id": cid, "reason": f"unknown action '{action}'"})
            continue
        if action == "delete":
            d = {"id": cid, "action": "delete"}
            if notes:
                d["notes"] = notes
                summary["with_notes"] += 1
            decisions.append(d)
            summary["deleted"] += 1
            continue
        if action == "merge":
            tid = ""
            if target in items and target != cid:
                tid = target
            elif target:
                hits = [x for x in by_name.get(target.casefold(), []) if x != cid]
                if len(hits) == 1:
                    tid = hits[0]
                elif len(hits) > 1:
                    ignored.append({"row": rownum, "id": cid, "reason": f"merge target '{target}' matches more than one code; use its id"})
                    continue
            if not tid:
                ignored.append({"row": rownum, "id": cid,
                                "reason": "merge target not found" if target else "merge without a 'merge into' value"})
                continue
            d = {"id": cid, "action": "merge", "merge_into": tid}
            if notes:
                d["notes"] = notes
                summary["with_notes"] += 1
            decisions.append(d)
            summary["merged"] += 1
            continue
        # keep / rename / blank: edits are found by comparing with the export
        d: dict = {"id": cid}
        if name and name != _norm(it.get("name")):
            d["name"] = name
            summary["renamed"] += 1
        if definition is not None and definition != _norm(it.get("definition")):
            d["definition"] = definition
            summary["redefined"] += 1
        if notes:
            d["notes"] = notes
            summary["with_notes"] += 1
        if "name" in d or "definition" in d:
            d["action"] = "rename"
            decisions.append(d)
        elif notes:
            d["action"] = "keep"
            decisions.append(d)
            summary["unchanged"] += 1
        else:
            summary["unchanged"] += 1
    summary["ignored"] = len(ignored)
    return {"kind": "code_review", "decisions": decisions, "additions": additions,
            "ignored": ignored, "summary": summary}


_YES = {"yes", "y", "true", "1", "x", "exclude", "excluded"}
_NO = {"", "no", "n", "false", "0", "keep", "include", "included"}


def _parse_papers(wb, payload: dict) -> dict:
    if "Papers" not in wb.sheetnames:
        raise SheetError("The workbook has no 'Papers' sheet.")
    ws = wb["Papers"]
    h = _headers(ws)
    if "source_id" not in h:
        raise SheetError("The Papers sheet has no 'source_id' column. Keep the exported columns.")
    fields: list[str] = payload.get("fields") or []
    labels: dict = payload.get("field_labels") or {}
    # a field may be headed by its key or its label
    col_of: dict[str, int] = {}
    for k in fields:
        for cand in (k, labels.get(k, k)):
            if _norm(cand).casefold() in h:
                col_of[k] = h[_norm(cand).casefold()]
                break
    rows_by_id = {r["source_id"]: r for r in (payload.get("rows") or []) if r.get("source_id")}
    out_rows: list[dict] = []
    ignored: list[dict] = []
    summary = {"edited": 0, "fields_edited": 0, "excluded": 0, "reincluded": 0,
               "unchanged": 0, "ignored": 0, "with_notes": 0}
    seen: set[str] = set()
    for rownum, row in _rows(ws):
        sid = _cell(row, h, "source_id")
        if not sid:
            continue
        if sid not in rows_by_id:
            ignored.append({"row": rownum, "id": sid, "reason": "not a paper of this checkpoint"})
            continue
        if sid in seen:
            ignored.append({"row": rownum, "id": sid, "reason": "duplicate row; first row used"})
            continue
        seen.add(sid)
        orig = rows_by_id[sid]
        patch: dict = {}
        label = _cell(row, h, "label")
        if label and label != _norm(orig.get("label")):
            patch["label"] = label
        if "citation" in h:
            cit = _cell(row, h, "citation")
            if cit != _norm(orig.get("citation")):
                patch["citation"] = cit
        if "cited work" in h:
            cw = _cell(row, h, "cited work")
            if cw != _norm(orig.get("cited_work")):
                patch["cited_work"] = cw
        for k, i in col_of.items():
            v = _norm(row[i]) if i < len(row) else ""
            if v != _norm((orig.get("fields") or {}).get(k)):
                patch[k] = v
        if "exclude" in h:
            ev = _cell(row, h, "exclude").casefold()
            if ev in _YES:
                want = True
            elif ev in _NO:
                want = False
            else:
                ignored.append({"row": rownum, "id": sid, "reason": f"'exclude' must be yes or no, not '{ev}'"})
                continue
            if want != bool(orig.get("excluded")):
                patch["exclude"] = want
                summary["excluded" if want else "reincluded"] += 1
        notes = _cell(row, h, "notes")
        if notes:
            patch["notes"] = notes
            summary["with_notes"] += 1
        n_fields = sum(1 for k in patch if k not in ("exclude", "notes"))
        if n_fields:
            summary["edited"] += 1
            summary["fields_edited"] += n_fields
        if patch and (n_fields or "exclude" in patch or notes):
            out_rows.append({"source_id": sid, **patch})
        if not n_fields and "exclude" not in patch:
            summary["unchanged"] += 1
    summary["ignored"] = len(ignored)
    return {"kind": "extraction_review", "rows": out_rows, "ignored": ignored, "summary": summary}


def safe_filename(*parts: str, ext: str = "xlsx") -> str:
    stem = "_".join(re.sub(r"[^A-Za-z0-9._-]+", "_", p).strip("_") for p in parts if p)
    return (stem or "checkpoint")[:120] + "." + ext
