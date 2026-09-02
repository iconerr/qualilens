# Copyright 2026 Ashita Aggarwal and Suraj Commuri
# SPDX-License-Identifier: Apache-2.0

"""The checkpoint round trip: a checkpoint out as a workbook, a workbook back
in as the decisions the review screen stages. Every rule in the spec has a
test here; the last two run the whole path through the API on a mocked run
and require the same database state as the equivalent button decisions."""

import io
import json
import time

import pytest
from openpyxl import Workbook, load_workbook
from starlette.testclient import TestClient

import app.db as db
import app.llm as llm
from app import checkpoint_sheets as cs
from app.main import app, SESSION_TOKEN

AUTH = {"X-QualiLens-Token": SESSION_TOKEN}
client = TestClient(app, base_url="http://127.0.0.1", headers=AUTH)

PAYLOAD = {"kind": "code_review", "stage": "open_code", "items": [
    {"id": "aaa111", "name": "valuing price transparency", "definition": "d1", "excerpt_count": 2,
     "sample_excerpts": [{"quote": "the price was transparent"}]},
    {"id": "bbb222", "name": "fearing hidden fees", "definition": "d2", "excerpt_count": 1, "sample_excerpts": []},
    {"id": "ccc333", "name": "experiencing slow support", "definition": "d3", "excerpt_count": 1, "sample_excerpts": []},
    {"id": "ddd444", "name": "Fearing Hidden Fees", "definition": "a near-duplicate", "excerpt_count": 1, "sample_excerpts": []},
]}
META = {"project_name": "P", "run_id": "r1", "checkpoint_id": "cp1", "title": "Review initial codes",
        "stage": "review_codes", "exported_at": "2026-09-02 10:00"}
EXTRACTION = {"kind": "extraction_review", "stage": "extract_field",
              "fields": ["aims", "method", "sample", "findings", "limitations"],
              "field_labels": {"aims": "Aims", "method": "Method", "sample": "Sample",
                               "findings": "Findings", "limitations": "Limitations"},
              "rows": [
                  {"source_id": "s1", "filename": "Alpha.pdf", "label": "Alpha, 2021", "citation": "Alpha (2021)",
                   "cited_work": "", "fields": {"aims": "a", "method": "m", "sample": "s", "findings": "f", "limitations": "l"},
                   "excluded": False},
                  {"source_id": "s2", "filename": "Beta.pdf", "label": "Beta, 2022", "citation": "Beta (2022)",
                   "cited_work": "", "fields": {"aims": "a2", "method": "m2", "sample": "s2", "findings": "f2", "limitations": "l2"},
                   "excluded": False}]}


def _codes_wb(edits=None):
    """Export the fixture and apply cell edits: {(row, col_letter): value}."""
    data = cs.export_workbook("code_review", PAYLOAD, META)
    wb = load_workbook(io.BytesIO(data))
    for (r, col), v in (edits or {}).items():
        wb["Codes"][f"{col}{r}"] = v
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- export

def test_export_shape_types_and_mark():
    data = cs.export_workbook("code_review", PAYLOAD, META,
                              excerpts=[{"code_id": "aaa111", "code": "valuing price transparency",
                                         "via": "", "source": "doc.txt", "quote": "q", "memo": "m"}])
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Codes", "Excerpts", "About"]
    ws = wb["Codes"]
    assert [c.value for c in ws[1]] == ["id", "name", "definition", "excerpts", "sample excerpt",
                                        "action", "merge into", "notes"]
    assert ws["A2"].value == "aaa111" and ws["D2"].value == "2"     # text, not a number
    assert ws["A2"].number_format == "@" and ws["G2"].number_format == "@"
    assert ws.freeze_panes == "A2"
    assert ws.data_validations.dataValidation[0].formula1 == '"keep,merge,delete"'
    assert wb["Excerpts"]["E2"].value == "q"
    assert wb.defined_names[cs.MARK].attr_text == '"cp1"'
    about = {r[0]: r[1] for r in wb["About"].iter_rows(values_only=True) if r[0]}
    assert about["checkpoint id"] == "cp1" and about["kind"] == "code_review"


def test_export_refuses_other_kinds():
    with pytest.raises(cs.SheetError):
        cs.export_workbook("core_review", {}, META)


# ---------------------------------------------------------------- parse: the rules

def test_no_change_when_nothing_edited():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb(), "cp1")
    assert r["decisions"] == [] and r["additions"] == [] and r["ignored"] == []
    assert r["summary"]["unchanged"] == 4


def test_edits_are_found_by_comparison_no_action_word_needed():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb({
        (2, "B"): "valuing transparent pricing",          # rename
        (3, "C"): "worry about undisclosed charges",      # redefine
        (4, "B"): "experiencing slow support  ",          # whitespace only: not an edit
        (5, "C"): "",                                     # blank definition = deliberate clearing
    }), "cp1")
    by = {d["id"]: d for d in r["decisions"]}
    assert by["aaa111"] == {"id": "aaa111", "name": "valuing transparent pricing", "action": "rename"}
    assert by["bbb222"] == {"id": "bbb222", "definition": "worry about undisclosed charges", "action": "rename"}
    assert "ccc333" not in by
    assert by["ddd444"] == {"id": "ddd444", "definition": "", "action": "rename"}
    assert r["summary"]["renamed"] == 1 and r["summary"]["redefined"] == 2


def test_blank_name_keeps_the_old_name():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb({(2, "B"): ""}), "cp1")
    assert r["decisions"] == []


def test_delete_and_merge_by_id_and_by_name():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb({
        (2, "F"): "delete",
        (4, "F"): "merge", (4, "G"): "bbb222",                 # by id
        (5, "F"): "Merge", (5, "G"): "  experiencing SLOW support ",  # by name, any case, padded
    }), "cp1")
    by = {d["id"]: d for d in r["decisions"]}
    assert by["aaa111"] == {"id": "aaa111", "action": "delete"}
    assert by["ccc333"] == {"id": "ccc333", "action": "merge", "merge_into": "bbb222"}
    assert by["ddd444"] == {"id": "ddd444", "action": "merge", "merge_into": "ccc333"}
    assert r["summary"]["deleted"] == 1 and r["summary"]["merged"] == 2


def test_merge_target_ambiguous_missing_or_self_is_reported():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb({
        (2, "F"): "merge", (2, "G"): "fearing hidden fees",   # matches bbb222 AND ddd444 by name
        (3, "F"): "merge", (3, "G"): "no such code",
        (4, "F"): "merge",                                    # no target at all
        (5, "F"): "merge", (5, "G"): "ddd444",                # itself
    }), "cp1")
    assert r["decisions"] == []
    reasons = {g["id"]: g["reason"] for g in r["ignored"]}
    assert "more than one code" in reasons["aaa111"]
    assert reasons["bbb222"] == "merge target not found"
    assert "without a 'merge into'" in reasons["ccc333"]
    assert reasons["ddd444"] == "merge target not found"


def test_unknown_action_unknown_id_duplicate_and_missing_row():
    data = cs.export_workbook("code_review", PAYLOAD, META)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Codes"]
    ws["F2"] = "archive"                       # not an action
    ws.delete_rows(3)                          # bbb222's row removed: no change to it
    ws.append(["zzz999", "stranger", "", "", "", "", "", ""])
    ws.append(["ccc333", "second row for ccc", "", "", "", "", "", ""])
    buf = io.BytesIO(); wb.save(buf)
    r = cs.parse_workbook("code_review", PAYLOAD, buf.getvalue(), "cp1")
    assert r["decisions"] == []
    reasons = {(g.get("id"), g["reason"]) for g in r["ignored"]}
    assert ("aaa111", "unknown action 'archive'") in reasons
    assert ("zzz999", "not a code of this checkpoint") in reasons
    assert any(i == "ccc333" and "duplicate" in why for i, why in reasons)
    assert r["summary"]["ignored"] == 3


def test_additions_and_notes():
    r = cs.parse_workbook("code_review", PAYLOAD, _codes_wb({
        (6, "B"): "brand new", (6, "C"): "added in Excel", (6, "H"): "seen in three interviews",
        (2, "H"): "kept on purpose",                       # a note with no other change
        (3, "F"): "delete", (3, "H"): "duplicate of ddd444",
    }), "cp1")
    assert r["additions"] == [{"name": "brand new", "definition": "added in Excel", "notes": "seen in three interviews"}]
    by = {d["id"]: d for d in r["decisions"]}
    assert by["aaa111"] == {"id": "aaa111", "action": "keep", "notes": "kept on purpose"}
    assert by["bbb222"] == {"id": "bbb222", "action": "delete", "notes": "duplicate of ddd444"}
    assert r["summary"]["with_notes"] == 3 and r["summary"]["added"] == 1


def test_wrong_checkpoint_unmarked_junk_and_oversize_are_refused():
    with pytest.raises(cs.SheetError, match="different checkpoint"):
        cs.parse_workbook("code_review", PAYLOAD, _codes_wb(), "cp-other")
    wb = Workbook(); wb.active.title = "Codes"; wb.active.append(["id", "name"])
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(cs.SheetError, match="no checkpoint mark"):
        cs.parse_workbook("code_review", PAYLOAD, buf.getvalue(), "cp1")
    with pytest.raises(cs.SheetError, match="not an Excel workbook"):
        cs.parse_workbook("code_review", PAYLOAD, b"PK\x03\x04 not really", "cp1")
    with pytest.raises(cs.SheetError, match="larger than"):
        cs.parse_workbook("code_review", PAYLOAD, b"x" * (cs.MAX_BYTES + 1), "cp1")


def test_missing_id_column_is_refused():
    data = cs.export_workbook("code_review", PAYLOAD, META)
    wb = load_workbook(io.BytesIO(data))
    wb["Codes"].delete_cols(1)
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(cs.SheetError, match="no 'id' column"):
        cs.parse_workbook("code_review", PAYLOAD, buf.getvalue(), "cp1")


def test_extraction_export_and_parse():
    meta = dict(META, checkpoint_id="cpx", title="Review the extraction table", stage="review_extraction")
    data = cs.export_workbook("extraction_review", EXTRACTION, meta,
                              excerpts=[{"source_id": "s1", "paper": "Alpha, 2021", "field": "Findings", "quote": "q", "memo": ""}])
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Papers", "Quotes", "About"]
    ws = wb["Papers"]
    heads = [c.value for c in ws[1]]
    assert heads == ["source_id", "filename", "label", "citation", "Aims", "Method", "Sample",
                     "Findings", "Limitations", "cited work", "exclude", "notes"]
    ws["H2"] = "Findings, corrected."          # a field edit
    ws["C2"] = ""                              # blank label: no change
    ws["K3"] = "yes"                           # exclude Beta
    ws["L3"] = "not a peer-reviewed source"
    ws.append(["s9", "Ghost.pdf", "Ghost", "", "", "", "", "", "", "", "", ""])
    buf = io.BytesIO(); wb.save(buf)
    r = cs.parse_workbook("extraction_review", EXTRACTION, buf.getvalue(), "cpx")
    rows = {x["source_id"]: x for x in r["rows"]}
    assert rows["s1"] == {"source_id": "s1", "findings": "Findings, corrected."}
    assert rows["s2"] == {"source_id": "s2", "exclude": True, "notes": "not a peer-reviewed source"}
    assert r["ignored"][0]["id"] == "s9"
    assert r["summary"]["edited"] == 1 and r["summary"]["excluded"] == 1 and r["summary"]["with_notes"] == 1
    ws["K3"] = "maybe"
    buf = io.BytesIO(); wb.save(buf)
    r = cs.parse_workbook("extraction_review", EXTRACTION, buf.getvalue(), "cpx")
    assert any("must be yes or no" in g["reason"] for g in r["ignored"])


# ---------------------------------------------------------------- through the API, on a mocked run

DOC_A = "We chose the vendor because the price was transparent. Hidden fees elsewhere made us nervous."
DOC_B = "Support was slow to respond at first. In the end the transparent pricing convinced our finance team."


def _ta_model(provider, model, api_key, system, user, max_tokens=8000, temperature=0.3):
    """A thematic-analysis model that answers from the prompt's stage markers."""
    import re
    usage = {"input_tokens": 10, "output_tokens": 5}
    if "reading data closely" in system:
        return json.dumps({"summary": "s", "memo": "m", "notable_features": []}), usage
    if "INITIAL CODES" in system:
        if "price was transparent" in user:
            return json.dumps({"codes": [
                {"name": "valuing price transparency", "definition": "d",
                 "excerpts": [{"quote": "the price was transparent", "memo": ""}]},
                {"name": "fearing hidden fees", "definition": "d",
                 "excerpts": [{"quote": "Hidden fees elsewhere made us nervous.", "memo": ""}]}]}), usage
        return json.dumps({"codes": [
            {"name": "valuing price transparency", "definition": "d",
             "excerpts": [{"quote": "the transparent pricing convinced our finance team", "memo": ""}]},
            {"name": "experiencing slow support", "definition": "d",
             "excerpts": [{"quote": "Support was slow to respond at first", "memo": ""}]}]}), usage
    if "CANDIDATE THEMES" in system:
        ids = re.findall(r"\[([0-9a-f]{12})\]", user)
        return json.dumps({"themes": [{"name": "Trust through transparency", "definition": "d",
                                       "rationale": "r", "code_ids": ids}]}), usage
    if "Phase 4" in system:
        ids = re.findall(r"\[([0-9a-f]{12})\]", user)
        return json.dumps({"reviews": [{"theme_id": i, "coherence": "strong", "distinctness": "adequate",
                                        "recommendation": "keep", "notes": "n"} for i in ids]}), usage
    if "Phase 5" in system:
        ids = re.findall(r"\[([0-9a-f]{12})\]", user)
        return json.dumps({"themes": [{"theme_id": i, "final_name": "Trust Through Transparency",
                                       "final_definition": "fd"} for i in ids]}), usage
    if "findings section" in system:
        return json.dumps({"sections": [{"heading": "Overview of Findings", "body": "b"}]}), usage
    return json.dumps({"ok": True}), usage


@pytest.fixture()
def mock_llm_ta(monkeypatch):
    monkeypatch.setattr(llm, "chat", _ta_model)
    client.put("/api/settings/keys", json={"anthropic": "sk-test"})
    conn = db.get_conn()
    conn.execute("UPDATE runs SET status='cancelled' WHERE status IN ('running','awaiting_review')")
    conn.commit()
    r = client.post("/api/projects", json={"name": "Sheet round trip", "method": "thematic",
                                           "config": {"provider": "anthropic", "research_question": "q"}})
    pid = r.json()["id"]
    for i, t in enumerate((DOC_A, DOC_B), 1):
        client.post(f"/api/projects/{pid}/sources",
                    files={"file": (f"doc_{i}.txt", io.BytesIO(t.encode()), "text/plain")}, data={"grp": ""})
    run_id = client.post(f"/api/projects/{pid}/runs").json()["run_id"]
    return pid, run_id


def _wait(run_id, *want, timeout=60):
    for _ in range(timeout * 20):
        d = client.get(f"/api/runs/{run_id}").json()
        if d["status"] in want:
            return d
        if d["status"] == "failed":
            raise AssertionError(f"run failed at {d['stage_name']}: {d['error']}")
        time.sleep(0.05)
    raise AssertionError(f"timeout waiting for {want}")


def test_round_trip_through_the_api_equals_button_decisions(mock_llm_ta):
    pid, run_id = mock_llm_ta
    d = _wait(run_id, "awaiting_review")
    cp = d["pending_checkpoint"]
    assert cp["payload"]["kind"] == "code_review"
    r = client.get(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet.xlsx")
    assert r.status_code == 200 and r.headers["content-disposition"].endswith('.xlsx"')
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Codes"]
    ids = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    names = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
    assert "valuing price transparency" in names
    i_val = names.index("valuing price transparency") + 2
    i_other = [i for i in range(2, ws.max_row + 1) if i != i_val][0]
    ws[f"B{i_val}"] = "valuing transparent pricing"; ws[f"H{i_val}"] = "gerund phrasing"
    ws[f"F{i_other}"] = "merge"; ws[f"G{i_other}"] = "valuing price transparency"   # old name still resolves
    ws.append(["", "from the sheet", "added", "", "", "", "", "co-coder suggestion"])
    buf = io.BytesIO(); wb.save(buf)
    r = client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet",
                    files={"file": ("codebook.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["summary"]["renamed"] == 1 and j["summary"]["merged"] == 1 and j["summary"]["added"] == 1
    assert j["imported_from"]["filename"] == "codebook.xlsx" and len(j["imported_from"]["sha256"]) == 64
    assert (db.UPLOADS_DIR / "checkpoints" / j["imported_from"]["stored"]).exists()
    # nothing applied yet: the checkpoint is still pending
    assert client.get(f"/api/runs/{run_id}").json()["status"] == "awaiting_review"
    # the same bytes again give the same decisions (stable ids, stable hash)
    r2 = client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet",
                     files={"file": ("again.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r2.json()["decisions"] == j["decisions"] and r2.json()["imported_from"]["sha256"] == j["imported_from"]["sha256"]
    # approve with what the screen would send
    r = client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/resolve",
                    json={"decisions": j["decisions"], "additions": j["additions"],
                          "stage": cp["payload"]["stage"], "imported_from": j["imported_from"]})
    assert r.status_code == 200, r.text
    conn = db.get_conn()
    rows = {r["name"]: dict(r) for r in conn.execute(
        "SELECT name, status, definition FROM codes WHERE run_id=? AND stage=?",
        (run_id, cp["payload"]["stage"])).fetchall()}
    assert rows["valuing transparent pricing"]["status"] == "active"
    assert rows["from the sheet"]["status"] == "active" and rows["from the sheet"]["definition"] == "added"
    merged_name = names[i_other - 2]
    assert rows[merged_name]["status"] == "merged"
    msgs = [e["message"] for e in client.get(f"/api/runs/{run_id}/events").json()]
    assert any("loaded from spreadsheet 'codebook.xlsx'" in m for m in msgs)
    assert any("Researcher note on code" in m and "gerund phrasing" in m for m in msgs)
    assert any("added code 'from the sheet' — note: co-coder suggestion" in m for m in msgs)
    stored = json.loads(conn.execute("SELECT resolution FROM checkpoints WHERE id=?", (cp["id"],)).fetchone()["resolution"])
    assert stored["imported_from"]["sha256"] == j["imported_from"]["sha256"]
    audit = client.get(f"/api/runs/{run_id}/audit.json").json()
    assert audit["checkpoints"][0]["resolution"]["imported_from"]["filename"] == "codebook.xlsx"
    d = _wait(run_id, "awaiting_review")
    client.post(f"/api/runs/{run_id}/checkpoints/{d['pending_checkpoint']['id']}/resolve", json={"decisions": []})
    _wait(run_id, "completed")


def test_sheet_endpoints_refuse_what_they_should(mock_llm_ta):
    pid, run_id = mock_llm_ta
    d = _wait(run_id, "awaiting_review")
    cp = d["pending_checkpoint"]
    assert client.get(f"/api/runs/{run_id}/checkpoints/nope/sheet.xlsx").status_code == 404
    r = client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet",
                    files={"file": ("x.xlsx", b"not a workbook", "application/octet-stream")})
    assert r.status_code == 400 and "not an Excel workbook" in r.json()["detail"]
    other = cs.export_workbook("code_review", PAYLOAD, META)     # marked for a different checkpoint
    r = client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet",
                    files={"file": ("other.xlsx", other, "application/octet-stream")})
    assert r.status_code == 400 and "different checkpoint" in r.json()["detail"]
    client.post(f"/api/runs/{run_id}/checkpoints/{cp['id']}/resolve", json={"decisions": []})
    r = client.get(f"/api/runs/{run_id}/checkpoints/{cp['id']}/sheet.xlsx")
    assert r.status_code == 400 and "already been resolved" in r.json()["detail"]
    d = _wait(run_id, "awaiting_review")
    client.post(f"/api/runs/{run_id}/checkpoints/{d['pending_checkpoint']['id']}/resolve", json={"decisions": []})
    _wait(run_id, "completed")
