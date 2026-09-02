# Copyright 2026 Ashita Aggarwal and Suraj Commuri
# SPDX-License-Identifier: Apache-2.0

"""The spreadsheet round trip through a real browser — optional, not CI.

Seeds a scratch database with a thematic run at its first checkpoint and a
literature-synthesis run at its extraction table (mock model, no keys),
serves the built interface, and drives the page: download the workbook,
edit it, upload it, check what is staged on screen, approve, and read the
audit trail. Needs Playwright with Chromium and openpyxl.

    QL_TREE=/path/to/QualiLens QL_PYTHON=/path/to/.venv/bin/python \
        python backend/tests/e2e_sheets_browser.py
"""

import io, json, os, pathlib, subprocess, sys, tempfile, time, urllib.request
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

TREE = os.environ.get("QL_TREE") or sys.exit("set QL_TREE"); PY = os.environ.get("QL_PYTHON", sys.executable); PORT = int(os.environ.get("QL_PORT", "8841")); ORIGIN = f"http://127.0.0.1:{PORT}"
DATA = pathlib.Path(tempfile.mkdtemp(prefix="ql_sheet_e2e_"))
OUT = os.environ.get("QL_SHOTS", tempfile.gettempdir())
fails = []
def check(c, m):
    print(("  ok  " if c else "  FAIL") + " " + m)
    if not c: fails.append(m)

# ---- seed: TA at review_codes, LS at review_extraction (mock model, in-process)
seed = r'''
import io, os, sys, json, time, pathlib
sys.path.insert(0, os.environ["QL_TREE"] + "/backend")
import app.db as db
DATA = pathlib.Path(os.environ["QUALILENS_DATA_DIR"]); db.DATA_DIR = DATA; db.DB_PATH = DATA/"qualilens.db"; db.UPLOADS_DIR = DATA/"uploads"; db.UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
os.environ["QUALILENS_TEST"] = "1"
src = open(os.environ["QL_TREE"] + "/backend/tests/e2e_methods.py").read().split("llm_mod.chat = fake_chat",1)[0]
src = src.replace('_db.DB_PATH = _pl.Path(_td) / "e2e.db"',"").replace('_db.UPLOADS_DIR = _pl.Path(_td) / "uploads"',"").replace("_db.UPLOADS_DIR.mkdir(exist_ok=True)","")
ns = {"__file__": os.environ["QL_TREE"] + "/backend/tests/e2e_methods.py"}; exec(compile(src,"h","exec"), ns)
import app.llm as llm; llm.chat = ns["fake_chat"]
from starlette.testclient import TestClient
from app.main import app, SESSION_TOKEN
c = TestClient(app, base_url="http://127.0.0.1", headers={"X-QualiLens-Token": SESSION_TOKEN})
c.put("/api/settings/keys", json={"anthropic": "sk-x"})
def mk(name, method, cfg, names):
    pid = c.post("/api/projects", json={"name": name, "method": method, "config": {"provider": "anthropic", **cfg}}).json()["id"]
    for t, fn in zip((ns["DOC_A"], ns["DOC_B"]), names):
        c.post(f"/api/projects/{pid}/sources", files={"file": (fn, io.BytesIO(t.encode()), "text/plain")}, data={"grp": ""})
    rid = c.post(f"/api/projects/{pid}/runs").json()["run_id"]
    for _ in range(600):
        d = c.get(f"/api/runs/{rid}").json()
        if d["status"] == "awaiting_review": return rid
        time.sleep(0.05)
    raise SystemExit("seed timeout")
ids = {"ta": mk("Vendor Selection Interviews", "thematic", {"research_question": "q"}, ("Interview 01.txt", "Interview 02.txt")),
       "ls": mk("Trust in Vendor Relationships", "literature_synthesis", {"research_question": "q"}, ("Alpha 2021.txt", "Beta 2022.txt"))}
(DATA/"ids.json").write_text(json.dumps(ids))
'''
env = dict(os.environ, QUALILENS_DATA_DIR=str(DATA), QL_TREE=TREE)
subprocess.run([PY, "-c", seed], env=env, check=True, capture_output=True)
ids = json.load(open(DATA / "ids.json"))

srv = subprocess.Popen([PY, "-m", "uvicorn", "app.main:app", "--host", "127.0.0.1", "--port", str(PORT)],
                       cwd=f"{TREE}/backend", env=env, stdout=open(DATA / "srv.log", "ab"), stderr=subprocess.STDOUT)
for _ in range(60):
    try: urllib.request.urlopen(ORIGIN + "/", timeout=1); break
    except Exception: time.sleep(0.25)


def wait_past(page, run_id, old_cp):
    page.wait_for_function(f"""async () => {{ const t = document.querySelector('meta[name=ql-token]').content;
        const r = await fetch('/api/runs/{run_id}', {{headers: {{'X-QualiLens-Token': t}}}}); const d = await r.json();
        return d.status !== 'awaiting_review' || (d.pending_checkpoint && d.pending_checkpoint.id !== '{old_cp}') }}""", timeout=60000)
    page.wait_for_timeout(500)
def cp_id(page, run_id):
    return page.evaluate(f"""async () => {{ const t = document.querySelector('meta[name=ql-token]').content;
        const r = await fetch('/api/runs/{run_id}', {{headers: {{'X-QualiLens-Token': t}}}}); const d = await r.json(); return d.pending_checkpoint && d.pending_checkpoint.id }}""")

with sync_playwright() as p:
    b = p.chromium.launch(args=["--no-sandbox"])
    ctx = b.new_context(viewport={"width": 1400, "height": 1000}, accept_downloads=True)
    page = ctx.new_page()
    page.on("dialog", lambda d: d.accept())

    # ================= code review =================
    page.goto(f"{ORIGIN}/#/runs/{ids['ta']}", wait_until="networkidle")
    page.wait_for_selector("text=Download as spreadsheet")
    with page.expect_download() as dl:
        page.click("text=Download as spreadsheet")
    import shutil
    path = DATA / "dl1.xlsx"; shutil.copy(dl.value.path(), path); name = dl.value.suggested_filename
    check(name.endswith(".xlsx") and "review_codes" in name, f"download named {name}")
    wb = load_workbook(path); ws = wb["Codes"]
    names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    check(len(names) == 3, f"sheet lists the 3 codes: {names}")
    # researcher edits: rename row 2 with a note; merge row 3 into row 2 by NAME; delete none; add one; one unknown id row
    ws["B2"] = "valuing transparent pricing"; ws["H2"] = "clearer as a gerund phrase"
    ws["F3"] = "merge"; ws["G3"] = "valuing price transparency"   # the ORIGINAL name still resolves
    ws["A5"] = ""; ws["B5"] = "sheet-added code"; ws["C5"] = "from Excel"
    ws["A6"] = "zzz999"; ws["B6"] = "stranger"
    edited = DATA / "codebook.xlsx"; wb.save(edited)
    page.set_input_files("input[type=file][accept='.xlsx']", str(edited))
    page.wait_for_selector("text=Loaded from", timeout=15000)
    line = page.inner_text(".sheetbar")
    check("1 renamed" in line and "1 merged" in line and "1 added" in line and "1 with notes" in line, f"summary: {line.strip()[:140]}")
    check("1 row ignored" in line, "the unknown-id row is reported as ignored")
    page.click("text=(show)")
    check("zzz999" in page.inner_text(".sheetbar") and "not a code of this checkpoint" in page.inner_text(".sheetbar"), "ignored row explained")
    # staged state on screen
    check(page.input_value(".code-item input[type=text]") == "valuing transparent pricing", "renamed name is in the input box")
    check("merging into" in page.inner_text(".review-layout"), "merge is staged and shown")
    check("clearer as a gerund phrase" in page.inner_text(".review-layout"), "note shown under the code")
    check(page.locator(".code-item input[placeholder='new code name']").input_value() == "sheet-added code", "added code staged")
    page.screenshot(path=f"{OUT}/9-sheet-staged.png")
    old = cp_id(page, ids['ta'])
    page.click("text=Approve & continue")
    wait_past(page, ids['ta'], old)
    events = page.evaluate(f"""async () => {{ const t = document.querySelector('meta[name=ql-token]').content;
        const r = await fetch('/api/runs/{ids['ta']}/events?after=0', {{headers: {{'X-QualiLens-Token': t}}}}); return (await r.json()).map(e => e.message) }}""")
    check(any("loaded from spreadsheet 'codebook.xlsx'" in m for m in events), "resolution event names the worksheet")
    check(any("Researcher note on code" in m and "gerund" in m for m in events), "note is in the audit trail")
    check(any("merged code" in m for m in events) and any("added code 'sheet-added code'" in m for m in events), "merge and addition applied")

    # ================= extraction review =================
    page.goto(f"{ORIGIN}/#/runs/{ids['ls']}", wait_until="networkidle")
    page.wait_for_selector("text=Download as spreadsheet")
    with page.expect_download() as dl:
        page.click("text=Download as spreadsheet")
    p2 = DATA / "dl2.xlsx"; shutil.copy(dl.value.path(), p2)
    wb = load_workbook(p2); ws = wb["Papers"]
    heads = [c.value for c in ws[1]]
    check(heads[:4] == ["source_id", "filename", "label", "citation"] and "Findings" in heads and "exclude" in heads, f"papers sheet headers {heads}")
    fi = heads.index("Findings") + 1; ex = heads.index("exclude") + 1; nt = heads.index("notes") + 1
    ws.cell(row=2, column=fi, value="Findings edited in the sheet.")
    ws.cell(row=2, column=nt, value="checked against the PDF")
    ws.cell(row=3, column=ex, value="yes")
    edited2 = DATA / "papers.xlsx"; wb.save(edited2)
    page.set_input_files("input[type=file][accept='.xlsx']", str(edited2))
    page.wait_for_selector("text=Loaded from", timeout=15000)
    line = page.inner_text(".sheetbar")
    check("1 papers edited" in line and "1 excluded" in line and "1 with notes" in line, f"extraction summary: {line.strip()[:140]}")
    check("excluded from synthesis" in page.inner_text("body"), "exclusion staged on screen")
    check("checked against the PDF" in page.inner_text("body"), "note shown under the paper")
    old = cp_id(page, ids['ls'])
    page.click("text=Approve & continue")
    wait_past(page, ids['ls'], old)
    events = page.evaluate(f"""async () => {{ const t = document.querySelector('meta[name=ql-token]').content;
        const r = await fetch('/api/runs/{ids['ls']}/events?after=0', {{headers: {{'X-QualiLens-Token': t}}}}); return (await r.json()).map(e => e.message) }}""")
    check(any("edited extraction 'findings'" in m for m in events), "findings edit applied")
    check(any("excluded" in m and "from the synthesis" in m for m in events), "exclusion applied")
    check(any("Researcher note on" in m and "PDF" in m for m in events), "extraction note in the audit trail")
    check(any("loaded from spreadsheet 'papers.xlsx'" in m for m in events), "resolution names the worksheet")
    b.close()
srv.terminate(); srv.wait()
print("ALL SHEET BROWSER CHECKS PASSED" if not fails else f"{len(fails)} FAILED")
sys.exit(1 if fails else 0)
